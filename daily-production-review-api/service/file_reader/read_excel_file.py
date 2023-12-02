import openpyxl
from openpyxl import load_workbook
import pandas as pd


def read_excel_file(file_path, sheet_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    table_html = df.to_html(index=False, escape=False)
    return table_html


def get_values_in_given_col(file_path, sheet_name, column_letter):
    try:
        wb = load_workbook(filename=file_path, data_only=True)
        sheet = wb[sheet_name]
        column_values = []

        for cell in sheet[column_letter]:
            column_values.append(cell.value)

        return column_values

    except Exception as e:
        print(f"An error occurred: {e}")
        return []


def find_index_in_list(column_values, target_value):
    try:
        index = column_values.index(target_value)
        return index
    except ValueError:
        return None


def read_excel_cell(file_path, sheet_name, cell_reference):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        cell_value = sheet[cell_reference].value
        workbook.close()
        return cell_value
    except Exception as e:
        return str(e)


def find_first_last_indices_mixed(sorted_list, target_value):
    first_index = None
    last_index = None

    for index, value in enumerate(sorted_list):
        if value == target_value:
            if first_index is None:
                first_index = index
            last_index = index

    if first_index is not None:
        return first_index, last_index
    else:
        return -1, -1


def get_values_from_given_row(excel_file_path, sheet_name, row_index):
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook[sheet_name]
    row = sheet[row_index]

    values = []
    columns_to_extract = [2, 5, 7, 9]

    for col_index in columns_to_extract:
        values.append(row[col_index].value)
    workbook.close()
    return values

def get_fields_for_a_date(excel_file_path, sheet_name):
    workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    sheet = workbook[sheet_name]
    values = []
    valueInCurrentRow = []
    column_indices = [2, 5, 7, 9]

    for row in sheet.iter_rows(min_row=4, values_only=True):
        valueInCurrentRow = []

        date_value = row[1]
        chassisNo_value = row[5]
        model_value = row[7]
        dealer = row[9]

        if date_value is not None and chassisNo_value is not None and model_value is not None and dealer is not None:
            valueInCurrentRow.append(date_value)
            valueInCurrentRow.append(chassisNo_value)
            valueInCurrentRow.append(model_value)
            valueInCurrentRow.append(dealer)
        if valueInCurrentRow:
            values.append(valueInCurrentRow)
    return values